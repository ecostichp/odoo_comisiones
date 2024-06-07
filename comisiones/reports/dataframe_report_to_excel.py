import pandas as pd
import numpy as np
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from typing import Literal

def dataframe_report_to_excel(
        df: pd.DataFrame,
        title: str,
        style_color: Literal["blue", "olive_green", "peach", "lilac", "mustard", "cherry"],
        excel_file: str=None,
        sheet_name: str=None,
        index_name: str=None,
        sep_gap: int=4,
        new_file_name: str = None
) -> None:
    """
    ## Convertir DataFrame a tabla de Excel
    Esta función convierte un Pandas DataFrame a una tabla estilizada estática de Excel
    y la coloca por debajo de las tablas existentes en un archivo provisto en los argumentos
    o crea un nuevo libro en caso de no ser provisto. 

    ### Forma de uso
    ```py
    dataframe_report_to_excel("Mi reporte", "cherry", df, "data/libro.xlsx", "Hoja 1")
    ```

    ### DataFrame
    Para ingresar el Pandas DataFrame, se coloca éste en el primer argumento posicional 

    ### Título de la tabla
    Para establecer el título de la tabla, se proporciona un valor de tipo `str` como segundo argumento posicional:
    ```py
    dataframe_report_to_excel("Mi reporte", ...)
    ```

    ### Color de la tabla
    Para establecer el color de la tabla, se proporciona alguna de las siguientes opciones de colores disponibles como tercer argumento posicional:
    - `blue`: Azul
    - `olive_green`: Verde oliva
    - `peach`: Durazno / Melón
    - `lilac`: Lila
    - `mustard`: Mostaza
    - `cherry`: Rojo cereza

    ### Archivo Excel
    Para espeficiar que la tabla se colocará en un archivo Excel existente, se provee en el parámetro `excel_file`:
    ```py
    dataframe_report_to_excel("Mi reporte", ..., excel_file= "data/libro.xlsx")
    ```

    En caso de no especificar un archivo, se creará uno nuevo con el nombre `reporte.xlsx`. Para cambiar el nombre del archivo
    resultante, leer la sección `Nombre del nuevo archivo`

    ### Hoja de Excel
    Para especificar una hoja en la cual se creará la tabla, se provee el parámetro `sheet_name`:
    ```py
    dataframe_report_to_excel("Mi reporte", ..., sheet_name="Hoja 1")
    ```

    - En caso de no especificar el nombre de la hoja, se creará una nueva con el nombre "Nueva hoja" y se guardará en el directorio actual
    del archivo en donde se ejecute esta función.
    - En caso de no encontrarse la hoja especificada por el usuario, se creará ésta con el nombre proporcionado.

    ### Separación entre tablas
    De haber tablas existentes en las hojas, se creará la tabla abajo de la última tabla existente en la hoja. El espacio de separación
    por defecto está establecido en 4 celdas. Para cambiar la cantidad de celdas de separación entre la última tabla existente en la hoja
    se provee el parámetro `sep_gap`.
    - La separación sólo aplica para la tabla actual que se va a escribir en el archivo Excel, no afecta la separación de las tablas existentes
    - De no haber tablas existentes en la hoja, la tabla se colocará directamente desde la fila 1 de la hoja y se ignorará por completo este parámetro.

    ### Nombre del nuevo archivo
    Para especificar el nombre del libro generado, se provee el parámetro `new_file_name`. Por defecto está establecido en `reporte.xlsx`.
    """
    # Parámetros iniciales de la función
    colors = {
        'blue': 'FF7D7DFF',
        'olive_green': 'FF88CB75',
        'peach': 'FFFFCA90',
        'lilac': 'FFBA8AFF',
        'mustard': 'FFEBCA4B',
        'cherry': 'FFFF405C'
    }

    xl_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    # Se carga o se crea el libro
    if excel_file:
        book = openpyxl.load_workbook(excel_file)
    else:
        book = Workbook()

    if not sheet_name:
        sheet_name = "Nueva hoja"

    # Se crea la hoja en el nuevo libro en caso de no haber sido proporcionada
    if not excel_file:
        book["Sheet"].title = sheet_name
    # Se crea la hoja en el libro existente en caso de no haber sido proporcionada
    else:
        book.create_sheet(sheet_name)
    
    # Se accede a la hoja o se crea una si el nombre de la hoja no se encontró
    try:
        sheet = book[sheet_name]
    except KeyError:
        book.create_sheet(sheet_name)
        sheet_name = book[sheet_name]

    # Se crea una copia del DataFrame
    df = df.copy()

    if index_name:
        df = df.reset_index(names=index_name)

    # Estilización de la tabla
    table_color = PatternFill("solid", fgColor=colors[style_color]) # Color de la tabla
    table_border = Border(
        top=Side(border_style="thin", color="FF000000"),
        left=Side(border_style="thin", color="FF000000"),
        right=Side(border_style="thin", color="FF000000"),
        bottom=Side(border_style="thin", color="FF000000"),
        # start=Side(border_style="double", color="FF000000"),
        # end=Side(border_style="double", color="FF000000"),
    ) # Borde de la tabla

    # Se establece el inicio horizontal en donde comenzará a colocarse la tabla
    start_row = sheet.max_row + sep_gap if sheet.max_row != 1 else 1

    # Se estiliza el borde de la celda
    sheet[f"A{start_row}"].border = table_border

    # Se crea el título de la tabla
    sheet.merge_cells(f"A{start_row}:{xl_columns[len(df.columns)-1]}{start_row}")

    # Se coloca el título en la celda combinada
    sheet[f"A{start_row}"] = title
    
    # Se coloca el título en la celda combinada
    sheet[f"A{start_row}"].alignment = Alignment(horizontal="center")

    # Se colorea la celda combinada del título
    sheet[f"A{start_row}"].fill = table_color

    # Se itera cada columna del DataFrame en conjunto con el índice de columnas de Excel
    for (df_column, xl_column) in zip(df.columns, xl_columns[:len(df.columns)]):

        # Valores para establecer el ancho de las columnas
        data = []

        # Se asignan los nombres de las columnas en las celdas
        sheet[f"{xl_column}{start_row+2}"] = df_column
        # Se colorea la celda
        sheet[f"{xl_column}{start_row+2}"].fill = table_color
        # Se da borde a la celda
        sheet[f"{xl_column}{start_row+2}"].border = table_border

        data.append(len(str(df_column)))


        # Se iteran los valores de cada columna del DataFrame y se reemplazan los valores np.nan por cadenas vacías
        for index, value in enumerate(df.replace({np.nan: ""})[df_column]):

            # Se colorea la celda si es el índice y si se especificó éste
            if index_name and xl_column == "A":
                sheet[f"{xl_column}{start_row+index+3}"].fill = table_color

            # Se asigna el valor de la celda del DataFrame en la celda correspondiente de Excel
            sheet[f"{xl_column}{start_row+index+3}"] = value
            # Se da borde a la celda
            sheet[f"{xl_column}{start_row+index+3}"].border = table_border
            # Se da formato de dos decilames a la celda
            sheet[f"{xl_column}{start_row+index+3}"].number_format = '#,##0.00'

            data.append(len(str(value)))

        column_width = sheet.column_dimensions[xl_column].width

        # Se establece el ancho de la columna
        sheet.column_dimensions[xl_column].width = np.array(data).max() * 1.2 if np.array(data).max() * 1.2 > column_width else column_width

    # Se crea el nombre del archivo a guardar en caso de no haber sido proporcionado
    if not new_file_name:
        new_file_name = "reporte.xlsx"

    book.save(new_file_name)